import os
import scipy.io
import numpy as np
import csv
from openpyxl import load_workbook

from PromptIQA.utils.dataset.loader_tools import *
from PromptIQA.utils.dataset.PromptIQADataset import PromptIQADataset

class LIVEC(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=True):
        super().__init__('livec')
        imgpath = scipy.io.loadmat(os.path.join(root, 'Data', 'AllImages_release.mat'))
        imgpath = imgpath['AllImages_release']
        imgpath = imgpath[7:1169]
        mos = scipy.io.loadmat(os.path.join(root, 'Data', 'AllMOS_release.mat'))
        labels = mos['AllMOS_release'].astype(np.float32)
        labels = labels[0][7:1169]

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, 'Images', imgpath[item][0][0]))
            gt.append(labels[item])
        # gt = normalization(gt)
        gt = list((np.array(gt) - 1) / 100)
        
        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
        
class AIGCIQA3W(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=True):
        super().__init__('aigciqa3w')

        imgname = []
        mos_all = []

        xls_file = os.path.join(root, 'info_train.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        prompt = []
        for row in rows:
            count += 1
            img_name = booksheet.cell(row=count, column=1).value
            imgname.append(img_name)
            mos = booksheet.cell(row=count, column=3).value
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            mos_all.append(mos)
            prompt.append(str(booksheet.cell(row=count, column=2).value))
            if count == 14002:
                break

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, 'train', imgname[item]))
            gt.append(mos_all[item])
        # gt = normalization(gt)
        gt = list(np.array(gt) / 1)
        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
    
class AIGCIQA2023(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=True):
        super().__init__('aigciqa2023')
        
        mos = scipy.io.loadmat(os.path.join(root, 'DATA', 'MOS', 'mosz1.mat'))
        labels = mos['MOSz'].astype(np.float32)

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, 'Image', 'allimg', f'{item}.png'))
            gt.append(labels[item][0])
        # gt = normalization(gt)
        gt = list(np.array(gt) / 100)
        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
    
class Koniq10k(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=True):
        super().__init__('koniq10k')
        
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'koniq10k_distributions_sets.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['MOS'])).astype(np.float32)
                mos_all.append(mos)

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, '1024x768', imgname[item]))
            gt.append(mos_all[item])
        # gt = normalization(gt)
        gt = list(np.array(gt) / 100)

        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size

class uhdiqa(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=True):
        super().__init__('udhiqa')
        
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'uhd-iqa-training-metadata.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['quality_mos'])).astype(np.float32)
                mos_all.append(mos)

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, 'challenge/training', imgname[item]))
            gt.append(mos_all[item])
        # gt = normalization(gt)

        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
        
class CGIQA6k(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=True):
        super().__init__('cgiqa6k')
        
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'mos.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['Image'])
                mos = np.array(float(row['MOS'])).astype(np.float32)
                mos_all.append(mos)

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, 'database', imgname[item]))
            gt.append(mos_all[item])
        gt = normalization(gt)

        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
    
class CSIQ(PromptIQADataset):
    def __init__(self, root, index, transform, patch_num=1, batch_size=11, istrain=True, dist_type=None):
        super().__init__('csiq')
        
        refpath = os.path.join(root, 'src_imgs')
        refname = getFileName(refpath, '.png')
        txtpath = os.path.join(root, 'csiq_label.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            if dist_type is None:
                imgnames.append((words[0]))
                target.append(words[1])
                ref_temp = words[0].split(".")
                refnames_all.append(ref_temp[0] + '.' + ref_temp[-1])
            else:
                if words[0].split('.')[1] == dist_type:
                    imgnames.append((words[0]))
                    target.append(words[1])
                    ref_temp = words[0].split(".")
                    refnames_all.append(ref_temp[0] + '.' + ref_temp[-1])

        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []
        gt = []

        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append(os.path.join(root, 'dst_imgs_all', imgnames[item]))
                    gt.append(labels[item])
        # gt = normalization(gt)
        gt = list(np.array(gt) / 1)
        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
        
        self.dist_type = dist_type
    
class TID2013Folder(PromptIQADataset):
    def __init__(self, root, index, transform, patch_num=1, batch_size=11, istrain=False):
        super().__init__('tid2013')
        
        refpath = os.path.join(root, 'reference_images')
        refname = getTIDFileName(refpath, '.bmp.BMP')
        txtpath = os.path.join(root, 'mos_with_names.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[1]))
            target.append(words[0])
            ref_temp = words[1].split("_")
            refnames_all.append(ref_temp[0][1:])
        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []
        gt = []
        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append(os.path.join(root, 'distorted_images', imgnames[item]))
                    gt.append(labels[item])
        # gt = normalization(gt)
        gt = list(np.array(gt) / 9)
        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
    
class KADID(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=True):
        super().__init__('kadid')
        
        imgpath = os.path.join(root, 'images')
        csv_file = os.path.join(root, 'dmos.csv')

        sample, gt = [], []
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                if (int(row['dist_img'].split('_')[0].replace('I', '')) - 1) in index:
                    sample.append(os.path.join(imgpath, row['dist_img']))
                    mos = np.array(float(row['dmos'])).astype(np.float32)
                    gt.append(mos)
        # gt = normalization(gt)
        gt = list((np.array(gt) - 1) / 5)

        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
  
class LIVEFolder(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=False):
        super().__init__('live')

        refpath = os.path.join(root, 'refimgs')
        refname = getFileName(refpath, '.bmp')

        jp2kroot = os.path.join(root, 'jp2k')
        jp2kname = self.getDistortionTypeFileName(jp2kroot, 227)

        jpegroot = os.path.join(root, 'jpeg')
        jpegname = self.getDistortionTypeFileName(jpegroot, 233)

        wnroot = os.path.join(root, 'wn')
        wnname = self.getDistortionTypeFileName(wnroot, 174)

        gblurroot = os.path.join(root, 'gblur')
        gblurname = self.getDistortionTypeFileName(gblurroot, 174)

        fastfadingroot = os.path.join(root, 'fastfading')
        fastfadingname = self.getDistortionTypeFileName(fastfadingroot, 174)

        imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname

        dmos = scipy.io.loadmat(os.path.join(root, 'dmos_realigned.mat'))
        labels = dmos['dmos'].astype(np.float32)

        orgs = dmos['orgs']
        refnames_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
        refnames_all = refnames_all['refnames_all']

        sample = []
        gt = []

        for i in range(0, len(index)):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = train_sel * ~orgs.astype(np.bool_)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[1].tolist()
            for j, item in enumerate(train_sel):
                sample.append(imgpath[item])
                gt.append(labels[0][item])
                
        # gt = normalization(gt)
        gt =list((np.array(gt) - 1) / 100)
        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
    
    def reshuffle(self):
        shuffle_sample, shuffle_gt = reshuffle(self.samples_p, self.gt_p)
        self.samples, self.gt = split_array(shuffle_sample, self.batch_size), split_array(shuffle_gt, self.batch_size)
        if len(self.samples[-1]) != self.batch_size:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]

class SPAQ(PromptIQADataset):
    def __init__(self, root, index, transform,  batch_size=11, istrain=False):
        super().__init__('spaq')
        
        sample = []
        gt = []
        xls_file = os.path.join(root, 'Annotations', 'MOS_and_Image_attribute_scores.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        
        for count, row in enumerate(rows, 2):
            if count - 2 in index:
                sample.append(os.path.join(root, 'img_resize', booksheet.cell(row=count, column=1).value))
                mos = booksheet.cell(row=count, column=2).value
                mos = np.array(mos)
                mos = mos.astype(np.float32)
                gt.append(mos)
            if count == 11126:
                break
        # gt = normalization(gt)
        gt = list(np.array(gt) / 100)
        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
        
class UWIQA(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=False):
        super().__init__('uwiqa')
        
        sample = []
        gt = []

        xls_file = os.path.join(root, 'IQA_Value.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        
        for count, row in enumerate(rows, 2):
            if count - 2 in index:
                sample.append(os.path.join(root, 'img', '{}.png'.format(str(booksheet.cell(row=count, column=1).value).zfill(4))))
                mos = booksheet.cell(row=count, column=2).value
                mos = np.array(mos)
                mos = mos.astype(np.float32)
                gt.append(mos)

        # gt = normalization(gt)
        gt = list(np.array(gt) / 1)
        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
    
class BID(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=False):
        super().__init__('bid')

        imgname = []
        mos_all = []

        xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for row in rows:
            count += 1
            img_num = booksheet.cell(row=count, column=1).value
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            imgname.append(img_name)
            mos = booksheet.cell(row=count, column=2).value
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            mos_all.append(mos)
            if count == 587:
                break

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, imgname[item]))
            gt.append(mos_all[item])
        # gt = normalization(gt)
        gt = list(np.array(gt) / 9)
        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
    
class GFIQA_20k(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=True):
        super().__init__('gfiqa_20k')
        
        imgname = []
        mos_all = []
        img_path = os.path.join(root, 'image')
        csv_file = os.path.join(root, 'mos.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['img_name'])
                mos = np.array(float(row['mos'])).astype(np.float32)
                mos_all.append(mos)

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(img_path, imgname[item]))
            gt.append(mos_all[item])
        # gt = normalization(gt)
        gt = list(np.array(gt) / 1)

        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size
    
class AGIQA_3k(PromptIQADataset):
    def __init__(self, root, index, transform, batch_size=11, istrain=True):
        super().__init__('agiqa_3k')
        
        imgname = []
        mos_all = []
        img_path = os.path.join(root, 'img')
        csv_file = os.path.join(root, 'data.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['name'])
                mos = np.array(float(row['mos_quality'])).astype(np.float32)
                mos_all.append(mos)

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(img_path, imgname[item]))
            gt.append(mos_all[item])
        # gt = normalization(gt)
        gt = list(np.array(gt) / 5)

        self.samples_p, self.gt_p = sample, gt

        self.samples, self.gt = split_array(sample, batch_size), split_array(gt, batch_size)
        if len(self.samples[-1]) != batch_size and istrain:
            self.samples = self.samples[:-1]
            self.gt = self.gt[:-1]
        self.transform = transform
        self.batch_size = batch_size