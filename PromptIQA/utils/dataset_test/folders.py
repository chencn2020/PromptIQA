import torch.utils.data as data
import torch

from PIL import Image
import os
import scipy.io
import numpy as np
import csv
from openpyxl import load_workbook
import cv2
import random

class LIVEC(data.Dataset):
    def __init__(self, root, index, transform):
        imgpath = scipy.io.loadmat(os.path.join(root, 'Data', 'AllImages_release.mat'))
        imgpath = imgpath['AllImages_release']
        imgpath = imgpath[7:1169]
        mos = scipy.io.loadmat(os.path.join(root, 'Data', 'AllMOS_release.mat'))
        labels = mos['AllMOS_release'].astype(np.float32)
        labels = labels[0][7:1169]

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, 'Images', imgpath[item][0][0]))
            gt.append(labels[item] / 100.)
        # gt = normalization(gt)

        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor, ''

    def __len__(self):
        length = len(self.samples)
        return length

    def get_promt(self, n=10):
        combined_data = list(zip(self.samples, self.gt))
        combined_data.sort(key=lambda x: x[1])

        values_to_insert = np.array([0.0, 1.0])
        position_to_insert = 0

        data_len = (self.__len__() - 2) // (n - 2)
        sample, gt = [], []
        sample.append(combined_data[0][0])
        gt.append(combined_data[0][1])
        
        for i in range(data_len, self.__len__(), data_len):
            sample.append(combined_data[i][0])
            gt.append(combined_data[i][1])
            if len(sample) == n - 1:
                break
        sample.append(combined_data[-1][0])
        gt.append(combined_data[-1][1])
        return prompt_data(sample, gt, self.transform)
        
class prompt_data(data.Dataset):
    def __init__(self, sample, gt, transform, div=1):
        self.samples, self.gt = sample, gt
        self.transform = transform
        self.div = div
        print('prompt GT is', gt)

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform, self.div)

        return img_tensor, gt_tensor

    def __len__(self):
        length = len(self.samples)
        return length


class Koniq10k(data.Dataset):
    def __init__(self, root, index, transform):
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'koniq10k_distributions_sets.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['MOS'])).astype(np.float32)
                mos_all.append(mos)

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, '1024x768', imgname[item]))
            gt.append(mos_all[item] / 100.)
        # gt = normalization(gt)

        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor, self.samples[index]

    def __len__(self):
        length = len(self.samples)
        return length

    def get_promt(self, n=10):
        combined_data = list(zip(self.samples, self.gt))
        combined_data.sort(key=lambda x: x[1])

        values_to_insert = np.array([0.0, 1.0])
        position_to_insert = 0

        data_len = (self.__len__() - 2) // (n - 2)
        sample, gt = [], []
        sample.append(combined_data[0][0])
        gt.append(combined_data[0][1])
        
        for i in range(data_len, self.__len__(), data_len):
            sample.append(combined_data[i][0])
            gt.append(combined_data[i][1])
            if len(sample) == n - 1:
                break
        sample.append(combined_data[-1][0])
        gt.append(combined_data[-1][1])
        return prompt_data(sample, gt, self.transform)

class CSIQ(data.Dataset):
    def __init__(self, root, index, transform, patch_num=1):

        refpath = os.path.join(root, 'src_imgs')
        refname = getFileName(refpath,'.png')
        txtpath = os.path.join(root, 'csiq_label.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[0]))
            target.append(words[1])
            ref_temp = words[0].split(".")
            refnames_all.append(ref_temp[0] + '.' + ref_temp[-1])

        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []
        gt = []

        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append(os.path.join(root, 'dst_imgs_all', imgnames[item]))
                    gt.append(labels[item])
        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor, ''

    def __len__(self):
        length = len(self.samples)
        return length

    def get_promt(self, n=10):
        combined_data = list(zip(self.samples, self.gt))
        combined_data.sort(key=lambda x: x[1])
        # random.shuffle(combined_data)

        values_to_insert = np.array([0.0, 1.0])
        position_to_insert = 0

        data_len = (self.__len__() - 2) // (n - 2)
        sample, gt = [], []
        sample.append(combined_data[0][0])
        gt.append(combined_data[0][1])
        
        for i in range(data_len, self.__len__(), data_len):
            sample.append(combined_data[i][0])
            gt.append(combined_data[i][1])
            if len(sample) == n - 1:
                break
        sample.append(combined_data[-1][0])
        gt.append(combined_data[-1][1])
        return prompt_data(sample, gt, self.transform)

class TID2013Folder(data.Dataset):

    def __init__(self, root, index, transform, patch_num=1):
        refpath = os.path.join(root, 'reference_images')
        refname = getTIDFileName(refpath,'.bmp.BMP')
        txtpath = os.path.join(root, 'mos_with_names.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[1]))
            target.append(words[0])
            ref_temp = words[1].split("_")
            refnames_all.append(ref_temp[0][1:])
        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []
        gt = []
        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append(os.path.join(root, 'distorted_images', imgnames[item]))
                    gt.append(labels[item] / 8)
        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor, ''

    def __len__(self):
        length = len(self.samples)
        return length

    def get_promt(self, n=10):
        combined_data = list(zip(self.samples, self.gt))
        combined_data.sort(key=lambda x: x[1])

        values_to_insert = np.array([0.0, 1.0])
        position_to_insert = 0

        data_len = (self.__len__() - 2) // (n - 2)
        sample, gt = [], []
        sample.append(combined_data[0][0])
        gt.append(combined_data[0][1])
        
        for i in range(data_len, self.__len__(), data_len):
            sample.append(combined_data[i][0])
            gt.append(combined_data[i][1])
            if len(sample) == n - 1:
                break
        sample.append(combined_data[-1][0])
        gt.append(combined_data[-1][1])
        return prompt_data(sample, gt, self.transform)
    

class KADID(data.Dataset):
    def __init__(self, root, index, transform):
        imgpath = os.path.join(root, 'images')
        
        csv_file = os.path.join(root, 'dmos.csv')
        
        imgname = []
        mos_all = []
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['dist_img'])
                mos = np.array(float(row['dmos'])).astype(np.float32)
                mos_all.append(mos)

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(imgpath, imgname[item]))
            gt.append(mos_all[item] / 5.)
        # gt = normalization(gt)
        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor, ''

    def __len__(self):
        length = len(self.samples)
        return length

    def get_promt(self, n=10):
        combined_data = list(zip(self.samples, self.gt))
        combined_data.sort(key=lambda x: x[1])

        values_to_insert = np.array([0.0, 1.0])
        position_to_insert = 0

        data_len = (self.__len__() - 2) // (n - 2)
        sample, gt = [], []
        sample.append(combined_data[0][0])
        gt.append(combined_data[0][1])
        
        for i in range(data_len, self.__len__(), data_len):
            sample.append(combined_data[i][0])
            gt.append(combined_data[i][1])
            if len(sample) == n - 1:
                break
        sample.append(combined_data[-1][0])
        gt.append(combined_data[-1][1])
        return prompt_data(sample, gt, self.transform)
    
class LIVEFolder(data.Dataset):

    def __init__(self, root, index, transform):

        refpath = os.path.join(root, 'refimgs')
        refname = getFileName(refpath, '.bmp')

        jp2kroot = os.path.join(root, 'jp2k')
        jp2kname = self.getDistortionTypeFileName(jp2kroot, 227)

        jpegroot = os.path.join(root, 'jpeg')
        jpegname = self.getDistortionTypeFileName(jpegroot, 233)

        wnroot = os.path.join(root, 'wn')
        wnname = self.getDistortionTypeFileName(wnroot, 174)

        gblurroot = os.path.join(root, 'gblur')
        gblurname = self.getDistortionTypeFileName(gblurroot, 174)

        fastfadingroot = os.path.join(root, 'fastfading')
        fastfadingname = self.getDistortionTypeFileName(fastfadingroot, 174)

        imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname

        dmos = scipy.io.loadmat(os.path.join(root, 'dmos_realigned.mat'))
        labels = dmos['dmos'].astype(np.float32)

        orgs = dmos['orgs']
        refnames_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
        refnames_all = refnames_all['refnames_all']

        sample = []
        gt = []

        for i in range(0, len(index)):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = train_sel * ~orgs.astype(np.bool_)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[1].tolist()
            for j, item in enumerate(train_sel):
                sample.append(imgpath[item])
                gt.append(labels[0][item] / 100.)
                # print(self.imgpath[item])
        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor, ''

    def __len__(self):
        length = len(self.samples)
        return length

    def getDistortionTypeFileName(self, path, num):
        filename = []
        index = 1
        for i in range(0, num):
            name = '%s%s%s' % ('img', str(index), '.bmp')
            filename.append(os.path.join(path, name))
            index = index + 1
        return filename
    def get_promt(self, n=10):
        combined_data = list(zip(self.samples, self.gt))
        combined_data.sort(key=lambda x: x[1])

        values_to_insert = np.array([0.0, 1.0])
        position_to_insert = 0

        data_len = (self.__len__() - 2) // (n - 2)
        sample, gt = [], []
        sample.append(combined_data[0][0])
        gt.append(combined_data[0][1])
        
        for i in range(data_len, self.__len__(), data_len):
            sample.append(combined_data[i][0])
            gt.append(combined_data[i][1])
            if len(sample) == n - 1:
                break
        sample.append(combined_data[-1][0])
        gt.append(combined_data[-1][1])
        return prompt_data(sample, gt, self.transform)
    
    
class FLIVE(data.Dataset):
    def __init__(self, root, index, transform):
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'labels_image.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['name'])
                mos = np.array(float(row['mos'])).astype(np.float32)
                mos_all.append(mos)

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, imgname[item]))
            gt.append(mos_all[item] / 100.)
        # gt = normalization(gt)

        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor, self.samples[index]

    def __len__(self):
        length = len(self.samples)
        return length

    def get_promt(self, n=10):
        combined_data = list(zip(self.samples, self.gt))
        combined_data.sort(key=lambda x: x[1])

        values_to_insert = np.array([0.0, 1.0])
        position_to_insert = 0

        data_len = (self.__len__() - 2) // (n - 2)
        sample, gt = [], []
        sample.append(combined_data[0][0])
        gt.append(combined_data[0][1])
        
        for i in range(data_len, self.__len__(), data_len):
            sample.append(combined_data[i][0])
            gt.append(combined_data[i][1])
            if len(sample) == n - 1:
                break
        sample.append(combined_data[-1][0])
        gt.append(combined_data[-1][1])
        return prompt_data(sample, gt, self.transform)

class SPAQ(data.Dataset):
    def __init__(self, root, index, transform, column=6):
        sample = []
        gt = []

        xls_file = os.path.join(root, 'Annotations', 'MOS_and_Image_attribute_scores.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        print('column', column)
        for count, row in enumerate(rows, 2):
            if count - 2 in index:
                sample.append(os.path.join(root, 'img_resize', booksheet.cell(row=count, column=1).value))
                mos = booksheet.cell(row=count, column=column).value
                mos = np.array(mos)
                mos = mos.astype(np.float32)
                gt.append(mos / 100.)
            if count == 11126:
                break
        # gt = normalization(gt)
        
        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor, self.samples[index]

    def __len__(self):
        length = len(self.samples)
        return length

    def get_promt(self, n=10):
        combined_data = list(zip(self.samples, self.gt))
        combined_data.sort(key=lambda x: x[1])

        values_to_insert = np.array([0.0, 1.0])
        position_to_insert = 0

        data_len = (self.__len__() - 2) // (n - 2)
        sample, gt = [], []
        sample.append(combined_data[0][0])
        gt.append(combined_data[0][1])
        
        for i in range(data_len, self.__len__(), data_len):
            sample.append(combined_data[i][0])
            gt.append(combined_data[i][1])
            if len(sample) == n - 1:
                break
        sample.append(combined_data[-1][0])
        gt.append(combined_data[-1][1])
        return prompt_data(sample, gt, self.transform)


class BID(data.Dataset):
    def __init__(self, root, index, transform):

        imgname = []
        mos_all = []

        xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for row in rows:
            count += 1
            img_num = booksheet.cell(row=count, column=1).value
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            imgname.append(img_name)
            mos = booksheet.cell(row=count, column=2).value
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            mos_all.append(mos / 5.)
            if count == 587:
                break

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, imgname[item]))
            gt.append(mos_all[item])

        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform, div=5)

        return img_tensor, gt_tensor, ""

    def __len__(self):
        length = len(self.samples)
        return length
    def get_promt(self, n=10):
        combined_data = list(zip(self.samples, self.gt))
        combined_data.sort(key=lambda x: x[1])

        data_len = (self.__len__() - 2) // (n - 2)
        sample, gt = [], []
        sample.append(combined_data[0][0])
        gt.append(combined_data[0][1])
        
        for i in range(data_len, self.__len__(), data_len):
            sample.append(combined_data[i][0])
            gt.append(combined_data[i][1])
            if len(sample) == n - 1:
                break
        sample.append(combined_data[-1][0])
        gt.append(combined_data[-1][1])
        return prompt_data(sample, gt, self.transform)

def get_item(samples, gt, index, transform, div=1):
    div = 1
    try:
        path, target = samples[index], gt[index]
        target = [target / div]

        values_to_insert = np.array([0.0, 1.0])
        position_to_insert = 0
        target = np.insert(target, position_to_insert, values_to_insert)

        sample = load_image(path)
        samples = {'img': sample, 'gt': target}
        samples = transform(samples)
    except:
        path, target = samples[0], gt[0]
        target = target / div

        values_to_insert = np.array([0.0, 1.0])
        position_to_insert = 0
        target = np.insert(target, position_to_insert, values_to_insert)

        sample = load_image(path)
        samples = {'img': sample, 'gt': target}
        samples = transform(samples)
        print('ERROR.')

    return samples['img'], samples['gt'].type(torch.FloatTensor)


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


def load_image(img_path):
    d_img = cv2.imread(img_path, cv2.IMREAD_COLOR)
    d_img = cv2.resize(d_img, (224, 224), interpolation=cv2.INTER_CUBIC)
    d_img = cv2.cvtColor(d_img, cv2.COLOR_BGR2RGB)
    d_img = np.array(d_img).astype('float32') / 255
    d_img = np.transpose(d_img, (2, 0, 1))
    
    return d_img

def normalization(data):
    data = np.array(data)
    range = np.max(data) - np.min(data)
    data = (data - np.min(data)) / range
    data = list(data.astype('float').reshape(-1, 1))

    return data

def getTIDFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if suffix.find(os.path.splitext(i)[1]) != -1:
            filename.append(i[1:3])
    return filename